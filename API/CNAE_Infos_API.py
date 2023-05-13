from flask import Flask, request, jsonify
import sqlite3
from openpyxl import load_workbook

app = Flask(__name__)
DATABASE_NAME = 'cnae_database.db'


def create_tables():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS cnae_infos 
                  (id INTEGER PRIMARY KEY AUTOINCREMENT,
                   codigo_seccao INTEGER,
                   seccao TEXT,
                   codigo_divisao INTEGER,
                   divisao TEXT,
                   codigo_grupo INTEGER,
                   grupo TEXT,
                   codigo_classe INTEGER,
                   classe TEXT,
                   codigo_subclasse INTEGER,
                   subclasse TEXT)''')

    conn.commit()
    conn.close()


@app.route('/insert_data', methods=['POST'])
def insert_data():
    file = request.files['file']
    wb = load_workbook(filename=file, read_only=True)
    ws = wb.active

    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    for row in ws.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT INTO cnae_infos (codigo_seccao, seccao, codigo_divisao, divisao, codigo_grupo, grupo, codigo_classe, classe, codigo_subclasse, subclasse) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", row)

    conn.commit()
    cursor.close()

    return 'Data inserted successfully!'

@app.route('/test', methods=['GET'])


@app.route('/view_data', methods=['GET'])
def select_data():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM cnae_infos")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    result = []
    for row in rows:
        data = {}
        data['id'] = row[0]
        data['codigo_seccao'] = row[1]
        data['seccao'] = row[2]
        data['codigo_divisao'] = row[3]
        data['divisao'] = row[4]
        data['codigo_grupo'] = row[5]
        data['grupo'] = row[6]
        data['codigo_classe'] = row[7]
        data['classe'] = row[8]
        data['codigo_subclasse'] = row[9]
        data['subclasse'] = row[10]
        result.append(data)

    return jsonify(result)


@app.route('/delete_data', methods=['DELETE'])
def delete_data():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM cnae_infos")
    conn.commit()
    cursor.close()
    conn.close()
    return 'All data deleted successfully!'



if __name__ == '__main__':
    create_tables()
    app.run(debug=False, port=3000)
